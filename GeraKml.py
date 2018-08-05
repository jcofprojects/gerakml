from pyautocad import Autocad, APoint
from openpyxl import load_workbook
import utm

class Coordenadas:
    def __init__(self, nome, pX, pY, pZ):
        self.nome = nome
        self.pX = pX
        self.pY = pY
        self.pZ = pZ

#Conexao com a Planilha de Coordenadas
wb = load_workbook(filename = 'Coordenadas.xlsx')
sheet_ranges = wb['Planilha1']
ws = wb.active   

print("GeraKml 1.0 - By Julio Felipe") 
tipoCoord = int(input("Projeção: 1 - UTM | 2 - LatLon : "))

#Criação do arquivo kml
file = open('GeraKmlFile.kml','w')
file.write('<?xml version="1.0" encoding="UTF-8"?>')
file.write('<kml xmlns="http://www.opengis.net/kml/2.2" xmlns:gx="http://www.google.com/kml/ext/2.2" xmlns:kml="http://www.opengis.net/kml/2.2" xmlns:atom="http://www.w3.org/2005/Atom">')
file.write('<Document>\n')

row_count = ws.max_row
for i in range(2, row_count - 1):
    cP1 = Coordenadas(sheet_ranges['A' + str(i)].value, sheet_ranges['B' + str(i)].value, sheet_ranges['C' + str(i)].value, sheet_ranges['D' + str(i)].value)
    print('{0}, {1}, {2}, {3}'.format(cP1.nome, cP1.pX, cP1.pY, cP1.pZ))
    file.write('<Placemark>')
    file.write('<name>' + cP1.nome + '</name>')
    file.write('<Point>')
    if (tipoCoord == 1):
        zoneNumber = int(input("Digite o numero da zona: "))
        zoneLetter = input("Digite a letra da zona: ")
        uLatLon = utm.to_latlon(cP1.pX, cP1.pY, zoneNumber, zoneLetter)
        file.write('<coordinates> {1}, {0}, {2}'.format(uLatLon[0], uLatLon[1], cP1.pZ) + '</coordinates>')
    else:
        file.write('<coordinates> {0}, {1}, {2}'.format(cP1.pX, cP1.pY, cP1.pZ) + '</coordinates>')
        file.write('</Point>')
        file.write('</Placemark>')

file.write('</Document>\n')
file.write('</kml>')
file.close() 
